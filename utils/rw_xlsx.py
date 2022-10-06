#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2021/5/13 
# @Author  : Mik


import openpyxl
from openpyxl.utils import get_column_letter

from config.base_path import base_path
from pathlib import Path


class ReadExcel:
    def __init__(self, sheet_name):
        """
        打开工作表
        :param sheet_name: 工作表名称
        """

        data_address = base_path + str(Path('/data', 'likeshop.xlsx'))
        workbook = openpyxl.load_workbook(data_address)
        # self.table = workbook.get_sheet_by_name[sheet_name]
        self.table = workbook[sheet_name]
        self.row_max1 = self.table.max_row
        self.col_max1 = self.table.max_column

    def get_sheet_info(self):
        """
        获取表信息
        :return: 名字，行数，列数
        """
        name = self.table.title
        row = self.row_max1
        col = self.col_max1
        return name, row, col

    def get_sheet_data_values(self):
        """
        获取整个表格的数据，返回列表嵌套字典格式，表头为key
        :return:
        """
        data_list = []
        for value in self.table.values:
            data_list.append(list(value))

    def get_sheet_data(self):
        """
        获取整个表格的数据，返回列表嵌套字典格式，表头为key
        :return:
        """
        data_list = []
        print(self.table.iter_rows())
        for row in self.table.iter_rows():
            temp_list = []
            for cell in row:
                temp_list.append(cell.value)
            data_list.append(temp_list)

        dict_list = []
        for i in range(1, len(data_list)):
            d = {}
            data = data_list[i]
            keys = data_list[0]
            for j in range(len(keys)):
                d[keys[j]] = data[j]
            dict_list.append(d)
        return dict_list

    def get_row(self, rowx):
        """
        获取一行的数据
        :param row: 传入行数从0开始
        :return: 这一行的数据
        """
        cellObj_list = []
        # row = list(self.table.rows)[rowx]
        row = self.table[rowx]
        for cellObj in row:
            cellObj_list.append(cellObj.value)
        return cellObj_list

    def get_rows(self, min, max):
        """
        获取指定行的数据
        :param min: 传入开始行，从0开始
        :param max: 传入结束行
        :return: 选择读取数据
        """

        rows = []
        for i in range(min, max + 1):
            rows.append(self.get_row(i))
        return rows

    def get_cols(self, min, max):
        """
        获取指定列的数据
        :param min: 传入开始列，从0开始
        :param max: 传入结束列
        :return: 选择读取数据
        """
        cols = []
        for i in range(min, max + 1):
            cols.append(self.get_row(i))
        return cols

    def get_col(self, colx):
        """
        获取一列的数据
        :param colx: 传入列数从0开始
        :return: 这一列的数据
        """
        cellObj_list = []
        loc = self.table[get_column_letter(colx)]
        for cellObj in loc:
            cellObj_list.append(cellObj.value)
        return cellObj_list

    def get_cell(self, cellx):
        """
        获取单元格数据
        :param cellx: 元组类型的行和列
        :return: 数据
        """
        cell = self.table.cell(*cellx).value
        return cell


class WriteExcel:
    def __init__(self, sheet='Sheet1'):
        """
        打开工作表
        :param sheet: 传入工作表名
        """
        self.data_address = base_path + str(Path('data', 'goodsname.xlsx'))
        self.book = openpyxl.load_workbook(self.data_address)
        self.bs = self.book[sheet]

    def write_call(self, cellx, val):
        """
        按单元格写入数据
        :param cellx: 传入单元格地址，例如：A1
        :param val: 传入写入值
        :return: None
        """
        self.bs[cellx] = val

    def write_row(self, val):
        """
        按行在末尾逐行写入数据
        :param val: 列表数据类型
        :return: None
        """
        self.bs.append(val)

    def save_book(self):
        self.book.save(self.data_address)


# e = ReadExcel()
# print(e.row_max1, type(e.row_max1))
# e1 = e.get_rows(0, e.row_max1)
# print(e1, type(e1))
#
# w=WriteExcel()
# w.write_call('A1','WWWWWWWWW')
# w.save_book()
if __name__ == '__main__':
    write_excel = WriteExcel()
    datas = [1, 2, 3]
    write_excel.write_row(datas)
    write_excel.save_book()
